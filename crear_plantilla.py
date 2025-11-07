"""
Script para crear la plantilla ligera de headers (solo filas 1-6).
Ejecutar UNA VEZ para generar plantilla/CARTERA_HEADERS.xlsx
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import os

def crear_plantilla_headers(ruta_machote, ruta_salida):
    """
    Extrae las filas 1-6 del machote y crea una plantilla ligera.
    
    Args:
        ruta_machote: Ruta al machote completo
        ruta_salida: Ruta donde guardar la plantilla ligera
    """
    print("=" * 80)
    print("CREANDO PLANTILLA LIGERA DE HEADERS")
    print("=" * 80)
    
    # 1. Cargar machote
    print(f"\n1. Cargando machote: {ruta_machote}")
    wb_machote = openpyxl.load_workbook(ruta_machote, data_only=True)
    
    # Buscar hoja CARTERA
    ws_machote = None
    for nombre in wb_machote.sheetnames:
        if nombre.upper() == 'CARTERA':
            ws_machote = wb_machote[nombre]
            break
    
    if not ws_machote:
        print("ERROR: No se encontró la hoja CARTERA")
        return False
    
    print(f"   Hoja CARTERA encontrada: {ws_machote.max_row} filas x {ws_machote.max_column} columnas")
    
    # 2. Crear nuevo workbook
    print(f"\n2. Creando workbook limpio...")
    wb_nuevo = Workbook()
    wb_nuevo.remove(wb_nuevo.active)
    ws_nuevo = wb_nuevo.create_sheet("CARTERA")
    
    # 3. Copiar filas 1-6 con todo el formato
    print(f"\n3. Copiando filas 1-6 con formato completo...")
    
    celdas_copiadas = 0
    for row_idx in range(1, 7):
        for col_idx in range(1, ws_machote.max_column + 1):
            celda_origen = ws_machote.cell(row_idx, col_idx)
            celda_destino = ws_nuevo.cell(row_idx, col_idx)
            
            # Copiar valor
            celda_destino.value = celda_origen.value
            
            # Copiar formato completo
            if celda_origen.has_style:
                celda_destino.font = copy(celda_origen.font)
                celda_destino.border = copy(celda_origen.border)
                celda_destino.fill = copy(celda_origen.fill)
                celda_destino.number_format = copy(celda_origen.number_format)
                celda_destino.protection = copy(celda_origen.protection)
                celda_destino.alignment = copy(celda_origen.alignment)
            
            if celda_origen.value is not None:
                celdas_copiadas += 1
    
    print(f"   Celdas con contenido copiadas: {celdas_copiadas}")
    
    # 4. Copiar anchos de columna
    print(f"\n4. Copiando dimensiones...")
    for col_idx in range(1, ws_machote.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_machote.column_dimensions:
            ws_nuevo.column_dimensions[col_letter].width = ws_machote.column_dimensions[col_letter].width
    
    # 5. Copiar alturas de fila
    for row_idx in range(1, 7):
        if row_idx in ws_machote.row_dimensions:
            ws_nuevo.row_dimensions[row_idx].height = ws_machote.row_dimensions[row_idx].height
    
    print(f"   Anchos y alturas copiados")
    
    # 6. Guardar plantilla
    print(f"\n5. Guardando plantilla: {ruta_salida}")
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb_nuevo.save(ruta_salida)
    
    # 7. Verificar tamaño
    tamaño_kb = os.path.getsize(ruta_salida) / 1024
    print(f"   Tamaño: {tamaño_kb:.1f} KB")
    
    print("\n" + "=" * 80)
    print("PLANTILLA CREADA EXITOSAMENTE")
    print("=" * 80)
    print(f"\nPlantilla: {ruta_salida}")
    print(f"Contenido: Filas 1-6 con formato completo")
    print(f"Tamaño: {tamaño_kb:.1f} KB (vs ~1.5 MB del machote)")
    print(f"\nEsta plantilla es independiente del machote y puede")
    print(f"distribuirse con el código sin problemas.")
    print("=" * 80)
    
    return True


if __name__ == '__main__':
    # Rutas
    MACHOTE = 'data/Copia de AntigüedadGrupal_machote.xlsm'
    PLANTILLA = 'plantilla/CARTERA_HEADERS.xlsx'
    
    # Crear plantilla
    exito = crear_plantilla_headers(MACHOTE, PLANTILLA)
    
    if exito:
        print("\nListo! Ahora formato_excel.py usara la plantilla ligera.")
    else:
        print("\nError al crear la plantilla.")

