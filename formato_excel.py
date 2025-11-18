"""
Módulo para guardar DataFrames en Excel con formato visual del machote.
Copia el formato exacto de la hoja CARTERA y convierte los datos en Tabla de Excel con totales automáticos.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.styles import Font, Fill, Border, Alignment, Protection, PatternFill
from copy import copy
import pandas as pd
import logging
import os

logger = logging.getLogger(__name__)


def copiar_headers(ws_origen, ws_destino):
    """
    Copia las filas 1-6 (headers) del machote al destino con formato completo.
    
    Args:
        ws_origen: Worksheet origen (machote CARTERA)
        ws_destino: Worksheet destino (nuevo archivo)
    """
    logger.info("Copiando headers (filas 1-6) con formato...")
    
    for row_idx in range(1, 7):
        for col_idx in range(1, ws_origen.max_column + 1):
            celda_origen = ws_origen.cell(row_idx, col_idx)
            celda_destino = ws_destino.cell(row_idx, col_idx)
            
            # Copiar valor
            celda_destino.value = celda_origen.value
            
            # Copiar formato completo
            if celda_origen.has_style:
                celda_destino.font = copy(celda_origen.font)
                celda_destino.border = copy(celda_origen.border)
                celda_destino.fill = copy(celda_origen.fill)
                celda_destino.number_format = copy(celda_origen.number_format)
                celda_destino.protection = copy(celda_origen.protection)
                celda_destino.alignment = copy(celda_origen.alignment)
    
    # Copiar anchos de columna
    for col_idx in range(1, ws_origen.max_column + 1):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_origen.column_dimensions:
            ws_destino.column_dimensions[col_letter].width = ws_origen.column_dimensions[col_letter].width
    
    # Copiar alturas de fila
    for row_idx in range(1, 7):
        if row_idx in ws_origen.row_dimensions:
            ws_destino.row_dimensions[row_idx].height = ws_origen.row_dimensions[row_idx].height
    
    logger.info(f"Headers copiados: {ws_origen.max_column} columnas")


def copiar_formato_fila(ws_origen, ws_destino, fila_origen, filas_destino):
    """
    Copia el formato de una fila del origen y lo replica a múltiples filas del destino.
    
    Args:
        ws_origen: Worksheet origen
        ws_destino: Worksheet destino
        fila_origen: Número de fila origen (ej. 7)
        filas_destino: Lista de números de fila destino (ej. [7, 8, 9, ...])
    """
    logger.info(f"Copiando formato de fila {fila_origen} a {len(filas_destino)} filas...")
    
    # Obtener número de columnas
    num_cols = ws_origen.max_column
    
    for fila_destino in filas_destino:
        for col_idx in range(1, num_cols + 1):
            celda_origen = ws_origen.cell(fila_origen, col_idx)
            celda_destino = ws_destino.cell(fila_destino, col_idx)
            
            # Copiar formato (NO valor)
            if celda_origen.has_style:
                celda_destino.font = copy(celda_origen.font)
                celda_destino.border = copy(celda_origen.border)
                celda_destino.fill = copy(celda_origen.fill)
                celda_destino.number_format = copy(celda_origen.number_format)
                celda_destino.protection = copy(celda_origen.protection)
                celda_destino.alignment = copy(celda_origen.alignment)
        
        # Copiar altura de fila
        if fila_origen in ws_origen.row_dimensions:
            ws_destino.row_dimensions[fila_destino].height = ws_origen.row_dimensions[fila_origen].height


def pegar_dataframe(ws, df, fila_inicio=7):
    """
    Pega los valores del DataFrame en el worksheet a partir de fila_inicio.
    
    Args:
        ws: Worksheet destino
        df: DataFrame con los datos
        fila_inicio: Fila donde empezar a pegar (default: 7)
    
    Returns:
        int: Última fila con datos
    """
    logger.info(f"Pegando DataFrame ({df.shape[0]} filas x {df.shape[1]} columnas) desde fila {fila_inicio}...")
    
    # Pegar datos fila por fila
    for idx, row in df.iterrows():
        fila_actual = fila_inicio + idx
        
        for col_idx, valor in enumerate(row, start=1):
            celda = ws.cell(fila_actual, col_idx)
            
            # Manejar valores NaN/None
            if pd.isna(valor):
                celda.value = None
            else:
                celda.value = valor
    
    ultima_fila = fila_inicio + len(df) - 1
    logger.info(f"Datos pegados hasta fila {ultima_fila}")
    
    return ultima_fila


def aplicar_formatos_columnas(ws, fila_inicio, fila_fin):
    """
    Aplica formatos de número a las columnas según especificación.
    
    Args:
        ws: Worksheet
        fila_inicio: Primera fila de datos (ej. 7)
        fila_fin: Última fila de datos
    """
    logger.info(f"Aplicando formatos de columnas desde fila {fila_inicio} hasta {fila_fin}...")
    
    # Formato de dinero: $#,##0.00
    formato_dinero = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    # Formato de dinero sin paréntesis (para negativos usa signo menos)
    # Usar el mismo formato que las otras columnas pero cambiar solo la sección de negativos
    # Formato estándar: '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    # Cambiar la segunda sección de: (#,##0.00) a: -#,##0.00 (sin paréntesis, con signo menos)
    formato_dinero_sin_parentesis = '_($* #,##0.00_);_($* -#,##0.00_);_($* "-"??_);_(@_)'
    
    # Formato de fecha corta: d/mm/aaaa
    formato_fecha_corta = 'd/mm/yyyy'
    
    # Formato de porcentaje: 0.00%
    formato_porcentaje = '0.00%'
    
    # Formato de texto (para mantener ceros a la izquierda)
    formato_texto = '@'
    
    # Aplicar formatos según columna (índice basado en 1)
    for fila in range(fila_inicio, fila_fin + 1):
        # E. Ciclo (columna 5): formato texto para mantener ceros a la izquierda
        ws.cell(fila, 5).number_format = formato_texto
        
        # F. Monto del crédito (columna 6): dinero
        ws.cell(fila, 6).number_format = formato_dinero
        
        # H. Fecha inicio del crédito (columna 8): fecha corta
        ws.cell(fila, 8).number_format = formato_fecha_corta
        
        # K. Hora de reunión (columna 11): ya está formateada como texto en el DataFrame
        
        # M. Pago semanal (columna 13): dinero
        ws.cell(fila, 13).number_format = formato_dinero
        
        # N. Próximo pago (columna 14): fecha corta
        ws.cell(fila, 14).number_format = formato_fecha_corta
        
        # O-V. Columnas de cartera (15-22): dinero
        for col in range(15, 23):
            if col == 18:  # Columna 18: Diferencia validación vigente (sin paréntesis)
                ws.cell(fila, col).number_format = formato_dinero_sin_parentesis
            else:
                ws.cell(fila, col).number_format = formato_dinero
        
        # W. %mora (columna 23): porcentaje
        ws.cell(fila, 23).number_format = formato_porcentaje
        
        # X-Z. Saldo en riesgo, saldo ahorro acumulado, monto promedio (24-26): dinero
        for col in range(24, 27):
            ws.cell(fila, col).number_format = formato_dinero
        
        # AG. Ahorro acumulado (columna 33): dinero
        ws.cell(fila, 33).number_format = formato_dinero
        
        # AH. %ahorro (columna 34): porcentaje
        ws.cell(fila, 34).number_format = formato_porcentaje
    
    logger.info(f"Formatos aplicados a {fila_fin - fila_inicio + 1} filas")


def crear_tabla_excel(ws, fila_inicio, fila_fin, num_cols, nombre_tabla="TablaCartera"):
    """
    Convierte el rango de datos en una Tabla de Excel con totales automáticos.
    
    Args:
        ws: Worksheet
        fila_inicio: Primera fila de headers de tabla (ej. 6)
        fila_fin: Última fila con datos
        num_cols: Número de columnas
        nombre_tabla: Nombre de la tabla (default: "TablaCartera")
    """
    logger.info(f"Creando tabla Excel '{nombre_tabla}' desde fila {fila_inicio} hasta {fila_fin}...")
    
    # Agregar una fila extra para totales
    fila_totales = fila_fin + 1
    
    # Definir rango de la tabla (incluye fila de totales)
    col_inicio = "A"
    col_fin = get_column_letter(num_cols)
    rango_tabla = f"{col_inicio}{fila_inicio}:{col_fin}{fila_totales}"
    
    logger.info(f"Rango de tabla: {rango_tabla}")
    logger.info(f"Fila de totales: {fila_totales}")
    
    # Obtener nombres de columnas desde la fila de headers (fila 6)
    nombres_columnas = []
    for col_idx in range(1, num_cols + 1):
        valor = ws.cell(fila_inicio, col_idx).value
        if valor:
            nombres_columnas.append(str(valor))
        else:
            nombres_columnas.append(f"Columna{col_idx}")
    
    logger.info(f"Nombres de columnas extraídos: {len(nombres_columnas)}")
    
    # Definir columnas con totales (índices basados en 0)
    columnas_con_totales = {
        5: "sum",   # Monto del crédito
        12: "sum",  # Pago semanal
        14: "sum",  # Cartera vigente sistema
        15: "sum",  # Cartera vigente inicial
        16: "sum",  # Cartera vigente calculada
        17: "sum",  # Cartera insoluta
        18: "sum",  # Diferencia validación vigente
        19: "sum",  # Ahorro consumido
        20: "sum",  # Cartera vencida estadística
        21: "sum",  # Cartera vencida total
        23: "sum",  # Saldo en riesgo
        24: "sum",  # Saldo ahorro acumulado
        25: "sum",  # Monto promedio del grupo
        32: "sum",  # Ahorro acumulado
    }
    
    # Crear columnas de tabla explícitamente
    table_columns = []
    for idx, nombre in enumerate(nombres_columnas):
        col_id = idx + 1
        
        if idx == 0:
            # Primera columna: label "Total"
            tc = TableColumn(id=col_id, name=nombre, totalsRowLabel="Total")
        elif idx in columnas_con_totales:
            # Columnas con suma
            tc = TableColumn(id=col_id, name=nombre, totalsRowFunction=columnas_con_totales[idx])
            logger.info(f"  Columna {idx + 1} '{nombre}': SUM")
        else:
            # Columnas sin totales
            tc = TableColumn(id=col_id, name=nombre)
        
        table_columns.append(tc)
    
    # Crear tabla con columnas configuradas
    tabla = Table(displayName=nombre_tabla, ref=rango_tabla, tableColumns=table_columns)
    
    # Aplicar estilo de tabla
    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo
    
    # Habilitar fila de totales
    tabla.totalsRowShown = True
    
    # Agregar tabla al worksheet
    ws.add_table(tabla)
    
    # Habilitar filtros automáticos en la fila de encabezados (fila 6)
    # Esto mostrará los iconos de filtro (triangulitos) en cada columna
    col_fin_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"{col_inicio}{fila_inicio}:{col_fin_letter}{fila_inicio}"
    logger.info(f"Filtros automáticos habilitados en rango: {ws.auto_filter.ref}")
    
    # Escribir fórmulas SUBTOTAL directamente en la fila de totales
    logger.info(f"Escribiendo fórmulas en fila de totales...")
    
    # Primera columna: texto "Total"
    ws.cell(fila_totales, 1).value = "Total"
    
    # Para cada columna con totales, escribir fórmula SUBTOTAL
    for col_idx, funcion in columnas_con_totales.items():
        col_letter = get_column_letter(col_idx + 1)
        # SUBTOTAL(109, ...) es la función SUM que ignora otras funciones SUBTOTAL
        # Rango desde fila inicio datos hasta última fila de datos
        formula = f"=SUBTOTAL(109,{col_letter}{fila_inicio + 1}:{col_letter}{fila_fin})"
        ws.cell(fila_totales, col_idx + 1).value = formula
        logger.info(f"  Fórmula en columna {col_idx + 1}: {formula}")
    
    logger.info(f"Tabla creada exitosamente con {fila_fin - fila_inicio} filas de datos + fila totales")
    logger.info(f"Totales configurados en {len(columnas_con_totales)} columnas con fórmulas SUBTOTAL")


def guardar_con_formato(df, ruta_plantilla, ruta_output):
    """
    Función principal: Guarda el DataFrame en Excel con formato de la plantilla.
    
    Args:
        df: DataFrame con los datos de CARTERA
        ruta_plantilla: Ruta a la plantilla de headers (plantilla/CARTERA_HEADERS.xlsx)
        ruta_output: Ruta del archivo de salida (.xlsx)
    
    Returns:
        str: Ruta del archivo generado
    """
    logger.info("=" * 80)
    logger.info("GUARDANDO DATAFRAME CON FORMATO")
    logger.info("=" * 80)
    
    # 1. Cargar plantilla ligera (solo headers, ~5KB)
    logger.info(f"\n1. Cargando plantilla de headers: {ruta_plantilla}")
    
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(
            f"Plantilla no encontrada: {ruta_plantilla}\n"
            f"Ejecuta 'python crear_plantilla.py' para generarla."
        )
    
    wb_plantilla = openpyxl.load_workbook(ruta_plantilla, data_only=True)
    ws_plantilla = wb_plantilla.active
    
    tamaño_kb = os.path.getsize(ruta_plantilla) / 1024
    logger.info(f"Plantilla cargada: {ws_plantilla.max_row} filas x {ws_plantilla.max_column} columnas ({tamaño_kb:.1f} KB)")
    
    # 2. Crear nuevo workbook
    logger.info("\n2. Creando nuevo workbook...")
    wb_nuevo = Workbook()
    wb_nuevo.remove(wb_nuevo.active)  # Eliminar hoja por defecto
    ws_nuevo = wb_nuevo.create_sheet("cartera")
    
    # 3. Copiar headers (filas 1-6) desde la plantilla
    logger.info("\n3. Copiando headers desde plantilla...")
    copiar_headers(ws_plantilla, ws_nuevo)
    
    # 3.1. Ajustar anchos de columnas específicas
    logger.info("\n3.1. Ajustando anchos de columnas específicas...")
    # Q (17): Cartera vigente calculada
    ws_nuevo.column_dimensions['Q'].width = 20.0
    # R (18): Cartera insoluta
    ws_nuevo.column_dimensions['R'].width = 20.0
    # S (19): Diferencia validación vigente
    ws_nuevo.column_dimensions['S'].width = 20.0
    # T (20): Ahorro consumido
    ws_nuevo.column_dimensions['T'].width = 20.0
    logger.info("Anchos ajustados: Q=20, R=20, S=20, T=20")
    
    # 4. Pegar datos del DataFrame
    logger.info("\n4. Pegando datos del DataFrame...")
    fila_inicio_datos = 7
    ultima_fila = pegar_dataframe(ws_nuevo, df, fila_inicio_datos)
    
    # 5. Aplicar formatos de columnas (dinero, fechas, porcentajes)
    logger.info("\n5. Aplicando formatos de columnas...")
    aplicar_formatos_columnas(ws_nuevo, fila_inicio_datos, ultima_fila)
    
    # 6. Crear tabla de Excel con totales
    logger.info("\n6. Creando tabla de Excel con totales...")
    try:
        crear_tabla_excel(
            ws_nuevo,
            fila_inicio=6,  # Headers de columna
            fila_fin=ultima_fila,
            num_cols=df.shape[1],
            nombre_tabla="TablaCartera"
        )
        
        # La fila de totales ya tiene formato por la tabla
        logger.info(f"Fila de totales: Formato aplicado automáticamente por la tabla")
        
        # 6.0. Re-aplicar formato a la columna 18 después de crear la tabla (la tabla puede sobrescribir formatos)
        # También aplicar a la fila de totales
        logger.info("\n6.0. Re-aplicando formato sin paréntesis a columna 18...")
        # Usar el mismo formato que las otras columnas pero sin paréntesis en negativos
        formato_dinero_sin_parentesis = '_($* #,##0.00_);_($* -#,##0.00_);_($* "-"??_);_(@_)'
        for fila in range(fila_inicio_datos, ultima_fila + 2):  # +2 para incluir fila de totales
            celda = ws_nuevo.cell(fila, 18)
            celda.number_format = formato_dinero_sin_parentesis
        logger.info(f"Formato re-aplicado a columna 18 (Diferencia validación vigente) desde fila {fila_inicio_datos} hasta {ultima_fila + 1}")
        
    except Exception as e:
        logger.warning(f"No se pudo crear tabla Excel: {e}")
        logger.warning("Continuando sin tabla (datos y formato están completos)")
    
    # 6.1. Congelar paneles para mantener encabezados visibles (filas 1-6)
    logger.info("\n6.1. Congelando paneles para mantener encabezados visibles...")
    ws_nuevo.freeze_panes = 'A7'  # Congela hasta la fila 6, fila 7 en adelante se desplaza
    logger.info("Paneles congelados: Filas 1-6 siempre visibles")
    
    # 7. Guardar archivo
    logger.info(f"\n7. Guardando archivo: {ruta_output}")
    wb_nuevo.save(ruta_output)
    
    logger.info("\n" + "=" * 80)
    logger.info("ARCHIVO GUARDADO EXITOSAMENTE")
    logger.info("=" * 80)
    logger.info(f"\nArchivo: {ruta_output}")
    logger.info(f"Filas de datos: {len(df)}")
    logger.info(f"Columnas: {df.shape[1]}")
    logger.info(f"Formato: Idéntico al machote")
    logger.info(f"Tabla Excel: Con totales automáticos y filtros")
    logger.info(f"Encabezados: Siempre visibles (paneles congelados)")
    logger.info("=" * 80)
    
    return ruta_output


def agregar_hoja_mora(ruta_output: str, df_mora: pd.DataFrame, ruta_plantilla: str):
    """
    Agrega la hoja MORA al archivo Excel existente con formato idéntico a CARTERA.
    
    Args:
        ruta_output: Ruta del archivo Excel a modificar
        df_mora: DataFrame con datos de MORA
        ruta_plantilla: Ruta de la plantilla con headers
    """
    logger.info("\n" + "=" * 80)
    logger.info("AGREGANDO HOJA MORA")
    logger.info("=" * 80)
    
    # 1. Abrir workbook existente
    logger.info(f"\n1. Abriendo archivo: {ruta_output}")
    wb = openpyxl.load_workbook(ruta_output)
    
    # 2. Crear nueva hoja "Mora"
    logger.info("\n2. Creando hoja 'Mora'")
    ws_mora = wb.create_sheet("Mora")
    
    # 3. Copiar headers de la plantilla (filas 1-6)
    logger.info("\n3. Copiando formato de headers desde plantilla")
    wb_plantilla = openpyxl.load_workbook(ruta_plantilla)
    ws_plantilla = wb_plantilla.active
    
    # Copiar solo primeras 14 columnas de las 6 filas
    for row_idx in range(1, 7):
        for col_idx in range(1, 15):  # Solo 14 columnas
            celda_origen = ws_plantilla.cell(row_idx, col_idx)
            celda_nueva = ws_mora.cell(row_idx, col_idx)
            
            celda_nueva.value = celda_origen.value
            if celda_origen.has_style:
                celda_nueva.font = copy(celda_origen.font)
                celda_nueva.border = copy(celda_origen.border)
                celda_nueva.fill = copy(celda_origen.fill)
                celda_nueva.number_format = copy(celda_origen.number_format)
                celda_nueva.alignment = copy(celda_origen.alignment)
    
    # Copiar dimensiones de columnas
    for col_idx in range(1, 15):
        col_letter = get_column_letter(col_idx)
        if col_letter in ws_plantilla.column_dimensions:
            ws_mora.column_dimensions[col_letter].width = ws_plantilla.column_dimensions[col_letter].width
    
    # 4. Sobrescribir headers con los nombres correctos (fila 6)
    logger.info("\n4. Estableciendo headers de MORA")
    headers_mora = [
        'Nombre del gerente',
        'Nombre del promotor',
        'ID GRUPO',
        'Nombre de grupo',
        'Ciclo',
        'Monto del crédito',
        'Semana',
        'Pago semanal',
        'Cartera vencida total',
        '%mora',
        'Saldo en riesgo',
        'Días de mora',
        'Mora potencial mensual',
        'Cartera vencida total'
    ]
    
    for col_idx, header in enumerate(headers_mora, start=1):
        ws_mora.cell(6, col_idx).value = header
    
    # 5. Pegar datos
    logger.info(f"\n5. Pegando {len(df_mora)} filas de datos")
    fila_inicio_datos = 7
    
    for row_idx, (_, row) in enumerate(df_mora.iterrows(), start=fila_inicio_datos):
        for col_idx, valor in enumerate(row, start=1):
            celda = ws_mora.cell(row_idx, col_idx)
            celda.value = valor
            
            # Aplicar formatos según tipo de columna
            if col_idx in [1, 2, 3, 4]:  # Texto
                celda.number_format = '@'
            elif col_idx == 5:  # Ciclo (número entero)
                celda.number_format = '0'
            elif col_idx in [6, 8, 9, 11, 13, 14]:  # Montos
                celda.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
            elif col_idx == 7:  # Semana (número entero)
                celda.number_format = '0'
            elif col_idx == 10:  # %mora (porcentaje)
                celda.number_format = '0.00%'
                # Fondo amarillo
                celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            elif col_idx == 12:  # Días de mora (número entero)
                celda.number_format = '0'
                # Fondo amarillo
                celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # 6. Crear tabla Excel
    logger.info("\n6. Creando tabla Excel con totales")
    
    if len(df_mora) > 0:
        ultima_fila = fila_inicio_datos + len(df_mora) - 1
        
        try:
            crear_tabla_mora(
                ws=ws_mora,
                fila_inicio=6,
                fila_fin=ultima_fila,
                num_cols=14,
                nombre_tabla="TablaMora"
            )
        except Exception as e:
            logger.warning(f"No se pudo crear tabla Excel: {e}")
    
    # 6.1. Congelar paneles para mantener encabezados visibles (filas 1-6)
    logger.info("\n6.1. Congelando paneles para mantener encabezados visibles...")
    ws_mora.freeze_panes = 'A7'  # Congela hasta la fila 6, fila 7 en adelante se desplaza
    logger.info("Paneles congelados: Filas 1-6 siempre visibles")
    
    # 7. Guardar
    logger.info(f"\n7. Guardando archivo con hoja MORA")
    wb.save(ruta_output)
    
    logger.info("\n" + "=" * 80)
    logger.info("HOJA MORA AGREGADA EXITOSAMENTE")
    logger.info("=" * 80)
    logger.info(f"\nRegistros en MORA: {len(df_mora)}")
    logger.info(f"Filtro: %mora > 5%")
    logger.info(f"Columnas amarillas: %mora, Días de mora")
    logger.info(f"Tabla Excel: Con totales automáticos y filtros")
    logger.info(f"Encabezados: Siempre visibles (paneles congelados)")
    logger.info("=" * 80)


def crear_tabla_mora(ws, fila_inicio, fila_fin, num_cols, nombre_tabla="TablaMora"):
    """
    Crea una tabla Excel en la hoja Mora con totales automáticos.
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
    
    # Definir rango de la tabla
    col_inicio = get_column_letter(1)
    col_fin = get_column_letter(num_cols)
    rango_tabla = f"{col_inicio}{fila_inicio}:{col_fin}{fila_fin + 1}"
    
    logger.info(f"   Creando tabla: {rango_tabla}")
    
    # Nombres de columnas
    nombres_columnas = [
        'Nombre del gerente',
        'Nombre del promotor',
        'ID GRUPO',
        'Nombre de grupo',
        'Ciclo',
        'Monto del crédito',
        'Semana',
        'Pago semanal',
        'Cartera vencida total',
        '%mora',
        'Saldo en riesgo',
        'Días de mora',
        'Mora potencial mensual',
        'Cartera vencida total'
    ]
    
    # Columnas con totales (índices basados en 0)
    columnas_con_totales = {
        5: "sum",   # Monto del crédito (columna F)
        7: "sum",   # Pago semanal (columna H)
        8: "sum",   # Cartera vencida total (columna I)
        10: "sum",  # Saldo en riesgo (columna K)
        12: "sum",  # Mora potencial mensual (columna M)
        13: "sum",  # Cartera vencida total calculada (columna N)
    }
    
    # Crear columnas de tabla
    table_columns = []
    for idx, nombre in enumerate(nombres_columnas):
        col_id = idx + 1
        if idx == 0:
            tc = TableColumn(id=col_id, name=nombre, totalsRowLabel="Total")
        elif idx in columnas_con_totales:
            tc = TableColumn(id=col_id, name=nombre, totalsRowFunction=columnas_con_totales[idx])
        else:
            tc = TableColumn(id=col_id, name=nombre)
        table_columns.append(tc)
    
    # Crear tabla
    tabla = Table(displayName=nombre_tabla, ref=rango_tabla, tableColumns=table_columns)
    tabla.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.totalsRowShown = True
    
    ws.add_table(tabla)
    
    # Habilitar filtros automáticos en la fila de encabezados (fila 6)
    # Esto mostrará los iconos de filtro (triangulitos) en cada columna
    col_inicio = get_column_letter(1)
    col_fin = get_column_letter(num_cols)
    ws.auto_filter.ref = f"{col_inicio}{fila_inicio}:{col_fin}{fila_inicio}"
    logger.info(f"   Tabla '{nombre_tabla}' creada con totales automáticos")
    logger.info(f"   Filtros automáticos habilitados en rango: {ws.auto_filter.ref}")
    
    # Escribir fórmulas SUBTOTAL en la fila de totales
    fila_totales = fila_fin + 1
    ws.cell(fila_totales, 1).value = "Total"
    
    for col_idx, funcion in columnas_con_totales.items():
        col_letter = get_column_letter(col_idx + 1)
        formula = f"=SUBTOTAL(109,{col_letter}{fila_inicio + 1}:{col_letter}{fila_fin})"
        ws.cell(fila_totales, col_idx + 1).value = formula
    
    logger.info(f"   Fórmulas SUBTOTAL escritas en {len(columnas_con_totales)} columnas")

