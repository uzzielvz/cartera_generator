import openpyxl
import os
import glob

# Buscar archivo machote
archivos = glob.glob('data/target/*machote*')
if not archivos:
    archivos = glob.glob('**/*machote*.xlsm', recursive=True)

print(f"Archivos encontrados: {archivos}")

if archivos:
    archivo = archivos[0]
    print(f"\nLeyendo: {archivo}")
    print(f"Archivo existe: {os.path.exists(archivo)}")
    
    try:
        wb = openpyxl.load_workbook(archivo, data_only=False, keep_vba=False)
        ws = wb['CARTERA']
        
        print(f"\n=== FÓRMULAS EN COLUMNA S (Diferencia validación vigente) ===")
        print(f"Fila totales: {ws.max_row}")
        
        # Leer fórmulas de algunas filas
        print("\nFórmulas en filas de datos:")
        for fila in range(7, 12):
            formula = ws.cell(fila, 19).value  # Columna S
            valor_data = None
            try:
                wb_data = openpyxl.load_workbook(archivo, data_only=True, keep_vba=False)
                ws_data = wb_data['CARTERA']
                valor_data = ws_data.cell(fila, 19).value
            except:
                pass
            print(f"  Fila {fila}: {formula} (valor: {valor_data})")
        
        # Leer fórmula del total
        print(f"\nFórmula del total (fila {ws.max_row}):")
        formula_total = ws.cell(ws.max_row, 19).value
        print(f"  Columna S: {formula_total}")
        
        # Leer fórmulas de otras columnas en el total
        print(f"\nFórmulas en fila de totales:")
        print(f"  Columna O (15): {ws.cell(ws.max_row, 15).value}")
        print(f"  Columna Q (17): {ws.cell(ws.max_row, 17).value}")
        print(f"  Columna S (19): {ws.cell(ws.max_row, 19).value}")
        
        # Leer valores calculados
        print(f"\nValores calculados (data_only=True):")
        try:
            wb_data = openpyxl.load_workbook(archivo, data_only=True, keep_vba=False)
            ws_data = wb_data['CARTERA']
            print(f"  Columna O total: {ws_data.cell(ws.max_row, 15).value}")
            print(f"  Columna Q total: {ws_data.cell(ws.max_row, 17).value}")
            print(f"  Columna S total: {ws_data.cell(ws.max_row, 19).value}")
        except Exception as e:
            print(f"  Error leyendo valores: {e}")
        
        # También verificar columnas O y Q para entender la fórmula
        print(f"\n=== COLUMNAS RELACIONADAS ===")
        print("Fila 7:")
        print(f"  Columna O (15): {ws.cell(7, 15).value} - Cartera vigente sistema")
        print(f"  Columna Q (17): {ws.cell(7, 17).value} - Cartera vigente calculada")
        print(f"  Columna S (19): {ws.cell(7, 19).value} - Diferencia validación vigente")
        
        # Buscar tablas
        print(f"\n=== TABLAS EN LA HOJA ===")
        print(f"Número de tablas: {len(ws.tables)}")
        if ws.tables:
            for nombre, tabla in ws.tables.items():
                print(f"  Tabla '{nombre}':")
                print(f"    Rango: {tabla.ref}")
                print(f"    Totales: {tabla.totalsRowShown}")
                if tabla.tableColumns:
                    print(f"    Columnas: {len(tabla.tableColumns)}")
                    # Buscar columna S (19)
                    for col in tabla.tableColumns:
                        if col.id == 19:
                            print(f"    Columna S (id={col.id}): totalsRowFunction={getattr(col, 'totalsRowFunction', 'None')}")
        
        # Buscar fórmulas SUBTOTAL
        print(f"\n=== FÓRMULAS SUBTOTAL ===")
        formulas_subtotal = []
        for i in range(1, min(ws.max_row + 1, 220)):
            for j in range(1, min(ws.max_column + 1, 40)):
                valor = ws.cell(i, j).value
                if isinstance(valor, str) and 'SUBTOTAL' in valor:
                    formulas_subtotal.append((i, j, valor))
        
        print(f"Fórmulas SUBTOTAL encontradas: {len(formulas_subtotal)}")
        for fila, col, formula in formulas_subtotal[:15]:
            print(f"  Fila {fila}, Col {col}: {formula}")
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
else:
    print("No se encontró archivo machote")

